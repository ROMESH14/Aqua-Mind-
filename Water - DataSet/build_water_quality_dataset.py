"""Build AquaMind water-quality Excel + CSV from the fish and plant catalogs.

Creates one species-range row for every fish and plant, then labelled
training readings (excellent / good / watch / critical) you can train on.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
FISH_XLSX = ROOT / "Fish - DataSet" / "AquaMind_Complete_Dataset.xlsx"
PLANT_XLSX = ROOT / "Plant - DataSet" / "AquaMind_Plant_Dataset_with_images.xlsx"
OUT_DIR = ROOT / "Water - DataSet"
OUT_XLSX = OUT_DIR / "AquaMind_Water_Quality_Dataset.xlsx"
OUT_RANGES_CSV = OUT_DIR / "species_water_ranges.csv"
OUT_TRAIN_CSV = OUT_DIR / "water_quality_training_samples.csv"

TEAL = "0F766E"
TEAL_DARK = "134E4A"
ROW_ALT = "F0FDFA"
WHITE = "FFFFFF"
AMBER = "FEF3C7"

QUALITY_LABELS = ("excellent", "good", "watch", "critical")


def _text(value, default=""):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    return str(value).strip()


def _num(value, default):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _round(value, digits):
    return round(float(value), digits)


def plant_chemistry(row):
    """Fill ammonia / nitrite / nitrate / DO for plants from care type."""
    care = _text(row.get("Care Level")).lower()
    growth = _text(row.get("Growth Rate")).lower()
    ptype = _text(row.get("Plant Type")).lower()
    name = _text(row.get("Common Name")).lower()

    sensitive = (
        "difficult" in care
        or "moss" in ptype
        or "carpet" in ptype
        or "baby tears" in name
        or "monte carlo" in name
        or "riccia" in name
    )
    floating = "floating" in ptype or any(
        key in name for key in ("duckweed", "frogbit", "lettuce", "hyacinth", "salvinia")
    )
    fast = "fast" in growth or "very fast" in growth

    if sensitive:
        max_nh3, max_no2 = 0.10, 0.25
        min_no3, ideal_no3, max_no3 = 5.0, 12.0, 25.0
        min_do, ideal_do = 5.0, 7.0
        sensitivity = "High"
    elif "moderate" in care:
        max_nh3, max_no2 = 0.20, 0.50
        min_no3, ideal_no3, max_no3 = 5.0, 15.0, 40.0
        min_do, ideal_do = 4.0, 6.5
        sensitivity = "Medium"
    else:
        max_nh3, max_no2 = 0.25, 0.75
        min_no3, ideal_no3, max_no3 = 5.0, 18.0, 45.0
        min_do, ideal_do = 4.0, 6.5
        sensitivity = "Low"

    if floating:
        max_nh3 = max(max_nh3, 0.30)
        max_no2 = max(max_no2, 0.75)
        max_no3 = max(max_no3, 60.0)
        ideal_no3 = max(ideal_no3, 20.0)
        min_do = min(min_do, 3.5)
    if fast and not sensitive:
        max_no3 = max(max_no3, 50.0)
        ideal_no3 = max(ideal_no3, 18.0)

    notes = []
    if floating or fast:
        notes.append("Fast grower / floater — useful nitrate sink.")
    if sensitive:
        notes.append("Sensitive foliage. Keep ammonia near 0 and nitrate moderate.")
    if not notes:
        notes.append("Hardy aquarium plant. Tolerates a wider chemistry window than most fish.")

    return {
        "max_ammonia": max_nh3,
        "max_nitrite": max_no2,
        "min_nitrate": min_no3,
        "ideal_nitrate": ideal_no3,
        "max_nitrate": max_no3,
        "min_do": min_do,
        "ideal_do": ideal_do,
        "sensitivity": sensitivity,
        "notes": " ".join(notes),
    }


def fish_sensitivity(row):
    care = _text(row.get("Care Level")).lower()
    ammonia = _num(row.get("Max Safe Ammonia (ppm)"), 0.02)
    nitrate = _num(row.get("Max Safe Nitrate (ppm)"), 20)
    name = _text(row.get("Common Name")).lower()
    if (
        "difficult" in care
        or ammonia <= 0.01
        or nitrate <= 15
        or "shrimp" in name
        or "stingray" in name
        or "discus" in name
        or "cardinal" in name
    ):
        return "High"
    if "moderate" in care:
        return "Medium"
    return "Low"


def fish_notes(row):
    name = _text(row.get("Common Name"))
    care = _text(row.get("Care Level"))
    ammonia = _num(row.get("Max Safe Ammonia (ppm)"), 0.02)
    if ammonia <= 0.01:
        return f"{name} is toxin-sensitive. Ammonia and nitrite must stay at 0 ppm."
    if "shrimp" in name.lower():
        return f"{name} needs stable, clean water. Copper and ammonia spikes are deadly."
    if care == "Difficult":
        return f"{name} needs tight water control. Test often and keep waste low."
    return f"{name} stays healthy when ammonia and nitrite are 0 and nitrate stays under the max."


def build_species_ranges():
    fish_df = pd.read_excel(FISH_XLSX, sheet_name="Complete Dataset")
    plant_df = pd.read_excel(PLANT_XLSX, sheet_name="Plant Species Data")
    rows = []

    for rec in fish_df.to_dict("records"):
        name = _text(rec.get("Common Name"))
        if not name:
            continue
        ph_min = _num(rec.get("Min pH"), 6.5)
        ph_max = _num(rec.get("Max pH"), 7.8)
        t_min = _num(rec.get("Min Temp (C)"), 22)
        t_max = _num(rec.get("Max Temp (C)"), 28)
        max_nh3 = _num(rec.get("Max Safe Ammonia (ppm)"), 0.02)
        max_no2 = _num(rec.get("Max Safe Nitrite (ppm)"), 0.02)
        max_no3 = _num(rec.get("Max Safe Nitrate (ppm)"), 20)
        min_do = _num(rec.get("Min Dissolved O2 (mg/L)"), 6.0)
        rows.append(
            {
                "ID": int(_num(rec.get("ID"), 0)),
                "Kind": "Fish",
                "Common Name": name,
                "Scientific Name": _text(rec.get("Scientific Name")),
                "Family": _text(rec.get("Family")),
                "Care Level": _text(rec.get("Care Level"), "Moderate"),
                "Sensitivity": fish_sensitivity(rec),
                "Min pH": _round(ph_min, 2),
                "Ideal pH": _round((ph_min + ph_max) / 2, 2),
                "Max pH": _round(ph_max, 2),
                "Min Temp (C)": _round(t_min, 1),
                "Ideal Temp (C)": _round((t_min + t_max) / 2, 1),
                "Max Temp (C)": _round(t_max, 1),
                "Min Ammonia (ppm)": 0.0,
                "Ideal Ammonia (ppm)": 0.0,
                "Max Safe Ammonia (ppm)": _round(max_nh3, 3),
                "Min Nitrite (ppm)": 0.0,
                "Ideal Nitrite (ppm)": 0.0,
                "Max Safe Nitrite (ppm)": _round(max_no2, 3),
                "Min Nitrate (ppm)": 0.0,
                "Ideal Nitrate (ppm)": _round(min(10.0, max_no3 * 0.4), 1),
                "Max Safe Nitrate (ppm)": _round(max_no3, 1),
                "Min Dissolved O2 (mg/L)": _round(min_do, 1),
                "Ideal Dissolved O2 (mg/L)": _round(min(10.0, min_do + 1.8), 1),
                "Max Dissolved O2 (mg/L)": 12.0,
                "Notes": fish_notes(rec),
                "Source": "Fish catalog",
            }
        )

    for rec in plant_df.to_dict("records"):
        name = _text(rec.get("Common Name"))
        if not name:
            continue
        chem = plant_chemistry(rec)
        ph_min = _num(rec.get("Min pH"), 6.0)
        ph_max = _num(rec.get("Max pH"), 7.5)
        t_min = _num(rec.get("Min Temp (C)"), 20)
        t_max = _num(rec.get("Max Temp (C)"), 28)
        rows.append(
            {
                "ID": 1000 + int(_num(rec.get("ID"), 0)),
                "Kind": "Plant",
                "Common Name": name,
                "Scientific Name": _text(rec.get("Scientific Name")),
                "Family": _text(rec.get("Family")),
                "Care Level": _text(rec.get("Care Level"), "Moderate"),
                "Sensitivity": chem["sensitivity"],
                "Min pH": _round(ph_min, 2),
                "Ideal pH": _round((ph_min + ph_max) / 2, 2),
                "Max pH": _round(ph_max, 2),
                "Min Temp (C)": _round(t_min, 1),
                "Ideal Temp (C)": _round((t_min + t_max) / 2, 1),
                "Max Temp (C)": _round(t_max, 1),
                "Min Ammonia (ppm)": 0.0,
                "Ideal Ammonia (ppm)": 0.0,
                "Max Safe Ammonia (ppm)": chem["max_ammonia"],
                "Min Nitrite (ppm)": 0.0,
                "Ideal Nitrite (ppm)": 0.0,
                "Max Safe Nitrite (ppm)": chem["max_nitrite"],
                "Min Nitrate (ppm)": chem["min_nitrate"],
                "Ideal Nitrate (ppm)": chem["ideal_nitrate"],
                "Max Safe Nitrate (ppm)": chem["max_nitrate"],
                "Min Dissolved O2 (mg/L)": chem["min_do"],
                "Ideal Dissolved O2 (mg/L)": chem["ideal_do"],
                "Max Dissolved O2 (mg/L)": 12.0,
                "Notes": chem["notes"],
                "Source": "Derived from plant catalog + aquarium chemistry rules",
            }
        )

    return pd.DataFrame(rows)


def _between(rng, low, high, pad=0.0):
    start, end = float(low) + pad, float(high) - pad
    if start >= end:
        return (float(low) + float(high)) / 2
    return float(rng.uniform(start, end))


def sample_excellent(rng, spec):
    return {
        "pH": _between(rng, spec["Min pH"], spec["Max pH"], 0.08),
        "Temperature": _between(rng, spec["Min Temp (C)"], spec["Max Temp (C)"], 0.6),
        "Ammonia": rng.uniform(0.0, max(0.001, spec["Max Safe Ammonia (ppm)"] * 0.20)),
        "Nitrite": rng.uniform(0.0, max(0.001, spec["Max Safe Nitrite (ppm)"] * 0.20)),
        "Nitrate": max(
            spec["Min Nitrate (ppm)"],
            rng.uniform(spec["Ideal Nitrate (ppm)"] * 0.4, spec["Ideal Nitrate (ppm)"] * 1.05),
        ),
        "DissolvedO2": rng.uniform(spec["Ideal Dissolved O2 (mg/L)"], spec["Max Dissolved O2 (mg/L)"] - 0.4),
    }


def sample_good(rng, spec):
    return {
        "pH": _between(rng, spec["Min pH"], spec["Max pH"], 0.02),
        "Temperature": _between(rng, spec["Min Temp (C)"], spec["Max Temp (C)"], 0.15),
        "Ammonia": rng.uniform(spec["Max Safe Ammonia (ppm)"] * 0.35, spec["Max Safe Ammonia (ppm)"] * 0.85),
        "Nitrite": rng.uniform(0.0, spec["Max Safe Nitrite (ppm)"] * 0.70),
        "Nitrate": rng.uniform(spec["Max Safe Nitrate (ppm)"] * 0.65, spec["Max Safe Nitrate (ppm)"] * 0.95),
        "DissolvedO2": rng.uniform(spec["Min Dissolved O2 (mg/L)"] + 0.2, spec["Ideal Dissolved O2 (mg/L)"]),
    }


def sample_watch(rng, spec):
    values = sample_excellent(rng, spec)
    kind = int(rng.integers(0, 6))
    if kind == 0:
        values["pH"] = spec["Max pH"] + rng.uniform(0.15, 0.55)
    elif kind == 1:
        values["pH"] = spec["Min pH"] - rng.uniform(0.15, 0.50)
    elif kind == 2:
        values["Temperature"] = spec["Max Temp (C)"] + rng.uniform(0.8, 2.8)
    elif kind == 3:
        values["Temperature"] = spec["Min Temp (C)"] - rng.uniform(0.8, 2.5)
    elif kind == 4:
        values["Nitrate"] = spec["Max Safe Nitrate (ppm)"] + rng.uniform(3, 14)
    else:
        values["Ammonia"] = spec["Max Safe Ammonia (ppm)"] + rng.uniform(0.005, 0.035)
        if values["Ammonia"] > max(0.05, spec["Max Safe Ammonia (ppm)"] * 2):
            values["Ammonia"] = spec["Max Safe Ammonia (ppm)"] + 0.03
    return values


def sample_critical(rng, spec):
    values = sample_excellent(rng, spec)
    kind = int(rng.integers(0, 4))
    if kind == 0:
        values["Ammonia"] = max(0.08, spec["Max Safe Ammonia (ppm)"] * 5) + rng.uniform(0.05, 0.9)
        values["Nitrite"] = max(0.25, spec["Max Safe Nitrite (ppm)"] * 6) + rng.uniform(0.1, 1.8)
    elif kind == 1:
        values["pH"] = float(rng.choice([rng.uniform(4.2, 5.3), rng.uniform(9.0, 10.4)]))
        values["Temperature"] = float(rng.choice([rng.uniform(12, 16), rng.uniform(33, 38)]))
    elif kind == 2:
        values["DissolvedO2"] = rng.uniform(1.4, max(2.1, spec["Min Dissolved O2 (mg/L)"] - 2.2))
        values["Ammonia"] = max(0.12, spec["Max Safe Ammonia (ppm)"] * 6)
    else:
        values["Nitrite"] = rng.uniform(1.0, 7.5)
        values["Nitrate"] = spec["Max Safe Nitrate (ppm)"] + rng.uniform(25, 80)
        values["Ammonia"] = rng.uniform(0.20, 1.40)
    return values


def violations(values, spec):
    issues = []
    if values["pH"] < spec["Min pH"] or values["pH"] > spec["Max pH"]:
        far = values["pH"] < spec["Min pH"] - 0.8 or values["pH"] > spec["Max pH"] + 0.8
        issues.append(("pH", "critical" if far else "watch"))
    if values["Temperature"] < spec["Min Temp (C)"] or values["Temperature"] > spec["Max Temp (C)"]:
        far = values["Temperature"] < spec["Min Temp (C)"] - 4 or values["Temperature"] > spec["Max Temp (C)"] + 4
        issues.append(("Temperature", "critical" if far else "watch"))
    if values["Ammonia"] > spec["Max Safe Ammonia (ppm)"]:
        far = values["Ammonia"] > max(0.05, spec["Max Safe Ammonia (ppm)"] * 2)
        issues.append(("Ammonia", "critical" if far else "watch"))
    if values["Nitrite"] > spec["Max Safe Nitrite (ppm)"]:
        far = values["Nitrite"] > max(0.25, spec["Max Safe Nitrite (ppm)"] * 3)
        issues.append(("Nitrite", "critical" if far else "watch"))
    if values["Nitrate"] > spec["Max Safe Nitrate (ppm)"]:
        issues.append(("Nitrate", "watch"))
    if values["DissolvedO2"] < spec["Min Dissolved O2 (mg/L)"]:
        far = values["DissolvedO2"] < spec["Min Dissolved O2 (mg/L)"] - 2
        issues.append(("DissolvedO2", "critical" if far else "watch"))
    return issues


def label_from_issues(issues, values, spec):
    if any(level == "critical" for _, level in issues) or len(issues) >= 3:
        return "critical"
    if issues:
        return "watch"
    if values["Ammonia"] > spec["Max Safe Ammonia (ppm)"] * 0.4:
        return "good"
    if values["Nitrate"] > spec["Max Safe Nitrate (ppm)"] * 0.7:
        return "good"
    return "excellent"


ACTIONS = {
    "pH": (
        "Correct pH slowly with small water changes only. Move no more than 0.2 pH per day.",
        "high",
    ),
    "Temperature": (
        "Adjust the heater 1°C at a time. Increase surface movement if the tank is hot.",
        "high",
    ),
    "Ammonia": (
        "Do a 40–50% water change with dechlorinated water. Pause feeding 24 hours and add beneficial bacteria.",
        "critical",
    ),
    "Nitrite": (
        "Change 40% of the water and add a nitrite detoxifier. Do not replace all filter media at once.",
        "critical",
    ),
    "Nitrate": (
        "Do a 25–40% water change, vacuum the substrate, and add fast-growing plants.",
        "medium",
    ),
    "DissolvedO2": (
        "Aim the filter at the surface or add an air stone. Fish gasping means act now.",
        "high",
    ),
}


def action_for(issues):
    if not issues:
        return "Keep weekly testing and regular water changes.", "low", ""
    primary = issues[0][0]
    detail, priority = ACTIONS[primary]
    if any(level == "critical" for _, level in issues):
        priority = "critical"
    return detail, priority, primary


def build_training_samples(ranges_df, rng):
    builders = (
        (sample_excellent, "excellent", 4),
        (sample_good, "good", 3),
        (sample_watch, "watch", 4),
        (sample_critical, "critical", 4),
    )
    rows = []
    sample_id = 1
    for spec in ranges_df.to_dict("records"):
        for builder, intended, count in builders:
            for _ in range(count):
                values = builder(rng, spec)
                issues = violations(values, spec)
                label = label_from_issues(issues, values, spec)
                if intended == "good" and label == "excellent" and rng.random() < 0.75:
                    label = "good"
                action, priority, primary = action_for(issues)
                rows.append(
                    {
                        "SampleID": sample_id,
                        "Species ID": spec["ID"],
                        "Kind": spec["Kind"],
                        "Common Name": spec["Common Name"],
                        "Scientific Name": spec["Scientific Name"],
                        "Sensitivity": spec["Sensitivity"],
                        "pH": _round(values["pH"], 2),
                        "Temperature": _round(values["Temperature"], 1),
                        "Ammonia": _round(values["Ammonia"], 3),
                        "Nitrite": _round(values["Nitrite"], 3),
                        "Nitrate": _round(values["Nitrate"], 1),
                        "DissolvedO2": _round(values["DissolvedO2"], 1),
                        "Quality": label,
                        "Failed Parameters": ", ".join(name for name, _ in issues) if issues else "",
                        "Primary Issue": primary,
                        "Recommended Action": action,
                        "Action Priority": priority,
                        "Min pH": spec["Min pH"],
                        "Max pH": spec["Max pH"],
                        "Min Temp (C)": spec["Min Temp (C)"],
                        "Max Temp (C)": spec["Max Temp (C)"],
                        "Max Safe Ammonia (ppm)": spec["Max Safe Ammonia (ppm)"],
                        "Max Safe Nitrite (ppm)": spec["Max Safe Nitrite (ppm)"],
                        "Max Safe Nitrate (ppm)": spec["Max Safe Nitrate (ppm)"],
                        "Min Dissolved O2 (mg/L)": spec["Min Dissolved O2 (mg/L)"],
                    }
                )
                sample_id += 1
    return pd.DataFrame(rows)


def action_guide_df():
    return pd.DataFrame(
        [
            {
                "Parameter": "pH Level",
                "Unit": "",
                "Typical safe community range": "6.5 – 7.8",
                "Watch when": "Outside the species min/max by up to 0.8",
                "Critical when": "More than 0.8 outside the species range",
                "What to do": ACTIONS["pH"][0],
            },
            {
                "Parameter": "Temperature",
                "Unit": "°C",
                "Typical safe community range": "22 – 28",
                "Watch when": "1–4°C outside the species range",
                "Critical when": "More than 4°C outside the species range",
                "What to do": ACTIONS["Temperature"][0],
            },
            {
                "Parameter": "Ammonia (NH3)",
                "Unit": "ppm",
                "Typical safe community range": "0 (max 0.02 for most fish)",
                "Watch when": "Just above the species max",
                "Critical when": ">0.05 ppm or more than 2× the species max",
                "What to do": ACTIONS["Ammonia"][0],
            },
            {
                "Parameter": "Nitrite (NO2)",
                "Unit": "ppm",
                "Typical safe community range": "0 (max 0.02 for most fish)",
                "Watch when": "Just above the species max",
                "Critical when": ">0.25 ppm or more than 3× the species max",
                "What to do": ACTIONS["Nitrite"][0],
            },
            {
                "Parameter": "Nitrate (NO3)",
                "Unit": "ppm",
                "Typical safe community range": "<20 for most fish; plants can use 10–40",
                "Watch when": "Above the species max",
                "Critical when": "Very high plus another toxin (ammonia or nitrite)",
                "What to do": ACTIONS["Nitrate"][0],
            },
            {
                "Parameter": "Dissolved O2",
                "Unit": "mg/L",
                "Typical safe community range": ">6 for fish; plants tolerate ~4+",
                "Watch when": "Just under the species minimum",
                "Critical when": "More than 2 mg/L below the minimum",
                "What to do": ACTIONS["DissolvedO2"][0],
            },
        ]
    )


def readme_rows(n_fish, n_plants, n_train):
    lines = [
        "AquaMind — Water Quality Dataset (Sri Lanka freshwater catalog)",
        "",
        "Use this file to train the Water Quality model. It matches the style of",
        "AquaMind_Plant_Dataset_with_images.xlsx and AquaMind_Complete_Dataset.xlsx.",
        "",
        f"Species ranges : {n_fish} fish + {n_plants} plants = {n_fish + n_plants} rows",
        f"Training samples: {n_train} labelled readings (excellent / good / watch / critical)",
        "",
        "HOW THE MODEL SHOULD WORK",
        "1. User picks a tank (its fish + plants).",
        "2. User logs the 6 parameters — typed in, or scanned from a test-kit photo.",
        "3. The model compares those readings to each species range in this file.",
        "4. It returns Quality + Recommended Action when water is watch or critical.",
        "",
        "SHEETS",
        "Species Water Ranges  — safe min / ideal / max for every fish and plant.",
        "Training Samples      — one labelled reading per row. Train on this sheet.",
        "Action Guide          — what to do when each parameter is bad.",
        "",
        "COLUMNS TO TRAIN ON",
        "Inputs : pH, Temperature, Ammonia, Nitrite, Nitrate, DissolvedO2",
        "         + the 8 range columns (Min pH … Min Dissolved O2)",
        "Target : Quality   (excellent, good, watch, critical)",
        "Extra  : Recommended Action, Primary Issue, Action Priority",
        "",
        "LABEL RULES",
        "excellent — all 6 values sit comfortably inside the species range.",
        "good      — still safe, but ammonia or nitrate is close to the limit.",
        "watch     — 1–2 values just outside the range.",
        "critical  — ammonia/nitrite spike, oxygen crash, extreme pH/temp, or 3+ faults.",
        "",
        "SOURCE",
        "Fish chemistry comes from AquaMind_Complete_Dataset.xlsx.",
        "Plant pH and temperature come from AquaMind_Plant_Dataset_with_images.xlsx.",
        "Plant ammonia, nitrite, nitrate and dissolved O2 are derived from care level,",
        "growth rate and plant type (plants tolerate more nitrate than fish).",
        "",
        "After you train, save the model (for example water_quality_classifier.pkl)",
        "and we will wire it into the Water page (manual log + image scan + result).",
    ]
    return pd.DataFrame({"AquaMind Water Quality Dataset": lines})


def style_sheet(ws, header=True, widths=None, wrap_cols=None):
    thin = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    header_fill = PatternFill("solid", fgColor=TEAL)
    header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    alt_fill = PatternFill("solid", fgColor=ROW_ALT)
    title_font = Font(name="Calibri", bold=True, color=TEAL_DARK, size=14)
    body_font = Font(name="Calibri", size=11)

    if header and ws.max_row >= 1:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin
        ws.row_dimensions[1].height = 32
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.font = body_font
                cell.border = thin
                cell.alignment = Alignment(vertical="center", wrap_text=cell.column in (wrap_cols or []))
                if cell.row % 2 == 0:
                    cell.fill = alt_fill
    else:
        if ws.max_row >= 1:
            ws["A1"].font = title_font
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(wrap_text=True, vertical="center")

    if widths:
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    else:
        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 18
    ws.sheet_properties.tabColor = TEAL


def write_excel(readme_df, ranges_df, train_df, guide_df):
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        readme_df.to_excel(writer, sheet_name="README - Read First", index=False)
        ranges_df.to_excel(writer, sheet_name="Species Water Ranges", index=False)
        train_df.to_excel(writer, sheet_name="Training Samples", index=False)
        guide_df.to_excel(writer, sheet_name="Action Guide", index=False)

    wb = load_workbook(OUT_XLSX)
    style_sheet(wb["README - Read First"], header=False, widths=[92])
    wb["README - Read First"]["A1"].fill = PatternFill("solid", fgColor=AMBER)
    for row in wb["README - Read First"].iter_rows(min_row=1, max_row=wb["README - Read First"].max_row):
        row[0].alignment = Alignment(wrap_text=True, vertical="center")
        wb["README - Read First"].row_dimensions[row[0].row].height = 18
    wb["README - Read First"].row_dimensions[1].height = 24

    style_sheet(
        wb["Species Water Ranges"],
        widths=[6, 8, 28, 28, 16, 12, 12, 10, 10, 10, 12, 13, 12, 14, 16, 18, 14, 14, 16, 14, 14, 16, 18, 16, 16, 55, 28],
        wrap_cols={26, 27},
    )
    style_sheet(
        wb["Training Samples"],
        widths=[10, 12, 8, 26, 26, 12, 8, 12, 10, 10, 10, 12, 12, 22, 14, 55, 14, 10, 10, 12, 12, 16, 16, 16, 16],
        wrap_cols={16},
    )
    style_sheet(wb["Action Guide"], widths=[18, 10, 38, 36, 42, 70], wrap_cols={3, 4, 5, 6})
    wb.save(OUT_XLSX)


def main():
    OUT_DIR.mkdir(exist_ok=True)
    ranges_df = build_species_ranges()
    rng = np.random.default_rng(42)
    train_df = build_training_samples(ranges_df, rng)
    guide_df = action_guide_df()
    n_fish = int((ranges_df["Kind"] == "Fish").sum())
    n_plants = int((ranges_df["Kind"] == "Plant").sum())
    readme_df = readme_rows(n_fish, n_plants, len(train_df))

    ranges_df.to_csv(OUT_RANGES_CSV, index=False)
    train_df.to_csv(OUT_TRAIN_CSV, index=False)
    write_excel(readme_df, ranges_df, train_df, guide_df)

    counts = train_df["Quality"].value_counts().to_dict()
    print(f"Species ranges : {n_fish} fish + {n_plants} plants")
    print(f"Training rows  : {len(train_df)}")
    print("Class mix      : " + ", ".join(f"{k}={counts.get(k, 0)}" for k in QUALITY_LABELS))
    print(f"Excel          : {OUT_XLSX}")
    print(f"Ranges CSV     : {OUT_RANGES_CSV}")
    print(f"Training CSV   : {OUT_TRAIN_CSV}")


if __name__ == "__main__":
    main()
